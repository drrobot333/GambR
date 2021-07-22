import openpyxl
import random

userEx = openpyxl.load_workbook('./user.xlsx')
sh = userEx.active

def loadFile():
    userEx = openpyxl.load_workbook('./user.xlsx')
    sh = userEx.active

def saveFile():
    userEx.save(filename='user.xlsx')
    userEx.close()

def findUser(userID):
    loadFile()
    i = 2
    while sh.cell(row=i, column = 1).value != None:
        if userID == sh.cell(row=i, column = 1).value:
            saveFile()
            return i
        i += 1


def findUserByReal(name):
    loadFile()
    i = 2
    while sh.cell(row=i, column = 6).value != None:
        if name == sh.cell(row=i, column = 6).value:
            saveFile()
            return i
        i += 1


def chkUser(userID):
    loadFile()
    i = 2
    while sh.cell(row=i, column = 1).value != None:
        if userID == sh.cell(row=i, column = 1).value:
            saveFile()
            return True
        i += 1

    else:
        saveFile()
        return False

def chkUserByReal(name):
    loadFile()
    i = 2
    while sh.cell(row=i, column = 6).value != None:
        if name == sh.cell(row=i, column = 6).value:
            saveFile()
            return True
        i += 1

    else:
        saveFile()
        return False
        

def addUser(ctx):
    loadFile()
    i = 2
    while sh.cell(row=i, column = 1).value != None:
        i += 1

    else:
        sh.cell(row=i, column = 1).value = str(ctx.author.id)
        sh.cell(row=i, column = 2).value = 10000
        sh.cell(row=i, column = 3).value = 0
        sh.cell(row=i, column = 4).value = 0
        sh.cell(row=i, column = 5).value = 0
        sh.cell(row=i, column = 6).value = ctx.author.name

    saveFile()
    
def batPossible(userID, money):
    loadFile()
    numID = findUser(userID)
    if sh.cell(row=numID, column = 2).value >= money and money > 0:
        return "MF"
        
    elif sh.cell(row=numID, column = 2).value < money:
        saveFile()
        return "MS"

    elif money <= 0:
        return "ME"
    
    saveFile()
    
def bat(userID, money):
    loadFile()
    numID = findUser(userID)
    if sh.cell(row=numID, column = 2).value >= money:
        sh.cell(row=numID, column = 2).value -= money
        sh.cell(row=numID, column = 5).value -= money
        saveFile()
        

def returnMoney(userID):
    loadFile()
    numID = findUser(userID)
    saveFile()
    return sh.cell(row=numID, column = 2).value

def loan(numID1, numID2, money):
    loadFile()
    
    user1Money = sh.cell(row=numID1, column = 2).value
    user2Money = sh.cell(row=numID2, column = 2).value

    if user1Money < money:
        saveFile()
        return "MS"

    else:
        sh.cell(row=numID2, column = 2).value = user2Money + money
        sh.cell(row=numID1, column = 2).value = user1Money - money
        saveFile()
        return "A"

def win(userID, money, value):
    loadFile()
    numID = findUser(userID)

    sh.cell(row=numID, column = 2).value += (money * value)
    sh.cell(row=numID, column = 4).value += 1
    sh.cell(row=numID, column = 5).value += (money * value)
    saveFile()

def lose(userID):
    loadFile()
    numID = findUser(userID)
    sh.cell(row=numID, column = 3).value += 1
    saveFile()

def info(userID):
    loadFile()
    numID = findUser(userID)
    money = sh.cell(row=numID, column = 2).value
    lost = sh.cell(row=numID, column = 3).value
    win = sh.cell(row=numID, column = 4).value
    profit = sh.cell(row=numID, column = 5).value
    saveFile()
    return [money, lost, win, profit]

def returnPeopleList(peopleList, newList):
    for i in range(len(peopleList)):
        if newList[0] == peopleList[i][0] and newList[2] == peopleList[i][2]:
            peopleList[i][-1] += newList[-1]
            return peopleList

    else:
        peopleList.append(newList)

    return peopleList

def returnWinner(matchInfo):
    name = matchInfo[:3]
    value = matchInfo[3:]
    if random.random() > 0.07: # 정배
        winner = random.choices(name, weights=[1/value[0], 1/value[1], 1/value[2]])[0]

    else: #역배
        winner = random.choices(name, weights=[value[0], value[1], value[2]])[0]

    return [winner, value[name.index(winner)]]

def returnResult(peopleList, matchInfo):
    #[Naussicaa, 456456456, 박지원 , 1000]
    winner, value = returnWinner(matchInfo)
    winList = []
    loseList = []

    for i in range(len(peopleList)):
        if peopleList[i][2] == winner:
            win(peopleList[i][1], peopleList[i][3], value)
            winList.append([peopleList[i][0], int(peopleList[i][3] * value)])

        else:
            lose(peopleList[i][1])
            loseList.append([peopleList[i][0], -peopleList[i][3]])

    return [winList, loseList, winner]

def returnStatus(peopleList, matchInfo):
    #[Naussicaa, 456456456, 박지원 , 1000]
    winList = []
    tieList = []
    loseList = []

    for i in range(len(peopleList)):
        if peopleList[i][2] == matchInfo[0]:
            winList.append([peopleList[i][0], peopleList[i][3]])

        elif peopleList[i][2] == matchInfo[1]:
            tieList.append([peopleList[i][0], peopleList[i][3]])

        else:
            lose(peopleList[i][1])
            loseList.append([peopleList[i][0], peopleList[i][3]])

    return [winList, tieList, loseList]
    
        
